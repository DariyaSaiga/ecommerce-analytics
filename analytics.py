import plotly.express as px
from config import run_query, load_queries
import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.formatting.rule import ColorScaleRule
from chart_create import (
    horizontal_bar_chart_revenue_by_state,
    line_chart_sales_by_month,
    bar_chart_payment_methods,
    histogram_delivery_by_category,
    pie_chart_reviews_by_category,
    scatter_top_products
)

queries = load_queries()

# --- Функции для графиков (с возвратом DataFrame) ---

def horizontal_bar_chart_revenue_by_state():
    df = run_query(queries["revenue_by_state"])
    fig = px.bar(
        df,
        x="total_revenue",
        y="customer_state",
        orientation="h",
        color="total_orders",
        title="Top States by Revenue"
    )
    fig.update_layout(xaxis_title="Revenue", yaxis_title="State")
    fig.show()
    return df

def line_chart_sales_by_month():
    df = run_query(queries["sales_by_month_states"])
    df["order_month"] = df["order_month"].dt.strftime("%Y-%m")
    fig = px.line(
        df,
        x="order_month",
        y="monthly_revenue",
        color="customer_state",
        markers=True,
        animation_frame="customer_state",
        title="Monthly Revenue by State (Interactive)"
    )
    fig.update_layout(xaxis_title="Month", yaxis_title="Revenue")
    fig.show()
    return df

def bar_chart_payment_methods():
    df = run_query(queries["payment_methods"])
    fig = px.bar(
        df,
        x="payment_type",
        y="transaction_count",
        color="total_processed",
        title="Transactions by Payment Method"
    )
    fig.update_layout(xaxis_title="Payment Type", yaxis_title="Number of Transactions")
    fig.show()
    return df

def histogram_delivery_by_category():
    df = run_query(queries["delivery_by_category"])
    fig = px.histogram(
        df,
        x="avg_delivery_days",
        color="product_category_name",
        nbins=15,
        title="Average Delivery Time by Category",
        barmode="overlay"
    )
    fig.update_layout(xaxis_title="Average Delivery Days", yaxis_title="Number of Orders")
    fig.show()
    return df

def pie_chart_reviews_by_category():
    df = run_query(queries["reviews_by_category"])
    df = df.sort_values(by="five_star_reviews", ascending=False).head(10)
    fig = px.pie(
        df,
        values="five_star_reviews",
        names="product_category_name",
        title="Proportion of Five-Star Reviews by Top Categories"
    )
    fig.show()
    return df

def scatter_top_products():
    df = run_query(queries["top_products"])
    fig = px.scatter(
        df,
        x="times_sold",
        y="total_revenue",
        size="avg_rating",
        color="product_category_name",
        hover_name="product_id",
        title="Top Products: Revenue vs Times Sold"
    )
    fig.show()
    return df

<<<<<<< HEAD
# --- Функция экспорта в Excel с форматированием ---

def export_to_excel(dataframes_dict, filename):
    filepath = f"exports/{filename}"
    with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
        for sheet_name, df in dataframes_dict.items():
            df.to_excel(writer, sheet_name=sheet_name, index=False)

    wb = load_workbook(filepath)
    total_rows = 0
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        total_rows += ws.max_row

        # Закрепляем первую строку
        ws.freeze_panes = "A2"
        # Фильтры
        ws.auto_filter.ref = ws.dimensions
        # Градиентная заливка числовых колонок
        for col in ws.iter_cols(min_row=2, max_row=ws.max_row):
            if all(isinstance(cell.value, (int, float)) or cell.value is None for cell in col):
                col_letter = col[0].column_letter
                rule = ColorScaleRule(
                    start_type="min", start_color="FFAA0000",
                    mid_type="percentile", mid_value=50, mid_color="FFFFFF00",
                    end_type="max", end_color="FF00AA00"
                )
                ws.conditional_formatting.add(f"{col_letter}2:{col_letter}{ws.max_row}", rule)

    wb.save(filepath)
    print(f"Создан файл {filename}, {len(wb.sheetnames)} листа(ов), {total_rows} строк")
=======
>>>>>>> 933bca9 (Добавлен экспорт дашборда Superset)

# --- Основной блок ---

if __name__ == "__main__":
    # Строим графики и получаем DataFrame
    df1 = horizontal_bar_chart_revenue_by_state()
    df2 = line_chart_sales_by_month() 
    df3 = bar_chart_payment_methods()
    df4 = histogram_delivery_by_category()
    df5 = pie_chart_reviews_by_category()
    df6 = scatter_top_products()

    # Словарь для экспорта
    dataframes_for_excel = {
        "Revenue by State": df1,
        "Monthly Sales": df2,
        "Payment Methods": df3,
        "Delivery by Category": df4,
        "Reviews by Category": df5,
        "Top Products": df6
    }

    # --- Функция экспорта в Excel с форматированием ---
    def export_to_excel(dataframes_dict, filename):
        with pd.ExcelWriter(f"exports/{filename}", engine='openpyxl') as writer:
            for sheet_name, df in dataframes_dict.items():
                df.to_excel(writer, sheet_name=sheet_name, index=False)

        wb = load_workbook(f"exports/{filename}")
        total_rows = 0
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            total_rows += ws.max_row

            # Закрепляем первую строку
            ws.freeze_panes = "A2"
            # Добавляем автофильтр
            ws.auto_filter.ref = ws.dimensions

            # Градиент для числовых колонок
            for col in ws.iter_cols(min_row=2, max_row=ws.max_row):
                if all(isinstance(cell.value, (int, float)) or cell.value is None for cell in col):
                    col_letter = col[0].column_letter
                    rule = ColorScaleRule(
                        start_type="min", start_color="FFAA0000",
                        mid_type="percentile", mid_value=50, mid_color="FFFFFF00",
                        end_type="max", end_color="FF00AA00"
                    )
                    ws.conditional_formatting.add(f"{col_letter}2:{col_letter}{ws.max_row}", rule)

        wb.save(f"exports/{filename}")
        print(f"Создан файл {filename}, {len(wb.sheetnames)} листа(ов), {total_rows} строк")

    # --- Экспортируем все таблицы ---
    export_to_excel(dataframes_for_excel, "sales_report.xlsx")